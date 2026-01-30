"""
Active Logging for CSES Workflow.

ARCHITECTURE:
- All data stored in JSON (source of truth)
- Rendered to Quarto Markdown after each change
- Atomic writes prevent corruption
- Never parse Markdown, only generate it

Files created:
- micro/.log_data.json (structured data)
- micro/{CODE}_{YEAR}_log.qmd (Quarto Markdown for humans)
"""

import json
import logging
import os
import tempfile
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from .state import WorkflowState

logger = logging.getLogger(__name__)


def atomic_write(path: Path, content: str):
    """
    Write content atomically using tempfile + os.replace.

    This ensures the file is never in a corrupted state:
    1. Write to temp file in same directory
    2. Atomic replace of target file
    3. Clean up on error
    """
    path = Path(path)
    dir_path = path.parent
    dir_path.mkdir(parents=True, exist_ok=True)

    fd, tmp_path = tempfile.mkstemp(dir=dir_path, suffix='.tmp')
    try:
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            f.write(content)
        os.replace(tmp_path, path)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise


@dataclass
class LogData:
    """
    Structured log data - stored as JSON, rendered to Markdown.

    This is the source of truth. The Markdown file is just a view.
    """
    meta: dict = field(default_factory=dict)
    deposited_files: dict = field(default_factory=dict)
    processing_notes: list = field(default_factory=list)
    collaborator_questions: list = field(default_factory=list)
    todo_items: list = field(default_factory=list)
    election_summary: str = ""
    study_design: dict = field(default_factory=dict)
    parties_leaders: str = ""

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> "LogData":
        return cls(
            meta=data.get("meta", {}),
            deposited_files=data.get("deposited_files", {}),
            processing_notes=data.get("processing_notes", []),
            collaborator_questions=data.get("collaborator_questions", []),
            todo_items=data.get("todo_items", []),
            election_summary=data.get("election_summary", ""),
            study_design=data.get("study_design", {}),
            parties_leaders=data.get("parties_leaders", "")
        )


class ActiveLogger:
    """
    Handles real-time logging with JSON storage and Markdown rendering.

    Data flow:
    1. Update operations modify self.log_data (dict operations - reliable)
    2. _save_and_render() saves JSON and regenerates Markdown
    3. Markdown is always complete and correct (generated, not edited)
    """

    def __init__(self, state: "WorkflowState"):
        """Initialize the active logger."""
        self.state = state
        self.working_dir = Path(state.working_dir) if state.working_dir else Path.cwd()

        # File paths
        self.json_path: Optional[Path] = None
        self.markdown_path: Optional[Path] = None

        # Log data (structured)
        self.log_data: Optional[LogData] = None

        # Initialize if state has country/year
        if state.country and state.year:
            self._initialize()

    def _initialize(self):
        """Initialize log data and files."""
        if not self.working_dir.exists():
            return

        country_code = self.state.country_code or self.state.country[:3].upper()
        year = self.state.year

        # Ensure micro folder exists
        micro_dir = self.working_dir / "micro"
        micro_dir.mkdir(parents=True, exist_ok=True)

        # File paths
        self.json_path = micro_dir / ".log_data.json"
        self.markdown_path = micro_dir / f"{country_code}_{year}_log.qmd"

        # Load or create log data
        if self.json_path.exists():
            self._load_log_data()
            print(f"Loaded log data from: micro/.log_data.json")
        else:
            self._create_default_log_data()
            self._save_and_render()
            print(f"Created log file: micro/{country_code}_{year}_log.qmd")

        # Update state with markdown path
        self.state.log_file = str(self.markdown_path)

    def _create_default_log_data(self):
        """Create default log data structure."""
        country_code = self.state.country_code or self.state.country[:3].upper()
        date_str = datetime.now().strftime("%Y-%m-%d")

        self.log_data = LogData(
            meta={
                "country": self.state.country,
                "country_code": country_code,
                "year": self.state.year,
                "processor": "[PROCESSOR NAME]",
                "created": date_str,
                "updated": date_str
            },
            deposited_files={
                "data_files": [],
                "questionnaires": [],
                "codebooks": [],
                "design_reports": [],
                "other": []
            },
            processing_notes=[],
            collaborator_questions=[],
            todo_items=[],
            election_summary="",
            study_design={
                "sample_design": "",
                "sample_size": "",
                "response_rate": "",
                "weighting": "",
                "collection_period": "",
                "mode": "",
                "field_lag": ""
            },
            parties_leaders=""
        )

    def _load_log_data(self):
        """Load log data from JSON."""
        try:
            data = json.loads(self.json_path.read_text(encoding='utf-8'))
            self.log_data = LogData.from_dict(data)
        except Exception as e:
            logger.error(f"Failed to load log data: {e}")
            self._create_default_log_data()

    def _save_and_render(self):
        """Save JSON and render Markdown atomically."""
        if not self.log_data or not self.json_path:
            return

        # Update timestamp
        self.log_data.meta["updated"] = datetime.now().strftime("%Y-%m-%d")

        # Save JSON
        json_content = json.dumps(self.log_data.to_dict(), indent=2, ensure_ascii=False)
        atomic_write(self.json_path, json_content)

        # Render and save Markdown
        markdown_content = self._render_markdown()
        atomic_write(self.markdown_path, markdown_content)

    def _render_markdown(self) -> str:
        """
        Render log data to Quarto Markdown.

        Follows CSES log file structure:
        1. Header (title, processor, country, year)
        2. Log File Instructions
        3. Log File Notes (deposited files, processing notes)
        4. Questions for Collaborator
        5. Things To Do Before Releasing
        6. Election Study Notes and Appendices
           - Election Summary
           - Study Design and Weights
           - Parties and Leaders
        """
        d = self.log_data
        m = d.meta
        code = m.get("country_code", "UNK")
        year = m.get("year", "0000")
        country = m.get("country", "Unknown")

        lines = [
            "---",
            f'title: "{code}_{year}_Mod6"',
            f'author: "{m.get("processor", "[PROCESSOR NAME]")}"',
            f'date: "{m.get("updated", "")}"',
            "format:",
            "  html:",
            "    toc: true",
            "    toc-depth: 3",
            "---",
            "",
            f"**Name:** {m.get('processor', '[PROCESSOR NAME]')}",
            f"",
            f"**Date:** {m.get('created', '')}",
            "",
            f"**Country:** {country}",
            "",
            f"**Election Year:** {year}",
            "",
            f"**Most recent update:** {m.get('updated', '')}",
            "",
            "---",
            "",
            "## Log File Instructions",
            "",
            "This Log File documents questions, issues, and challenges that arose while",
            "processing the Individual Datasets for inclusion in CSES Module 6.",
            "",
            "**Navigation:**",
            "",
            f"- [Log File Notes](#{code.lower()}-{year}-log-file-notes)",
            f"- [Questions for Collaborator](#{code.lower()}-{year}-questions)",
            f"- [Things To Do Before Releasing](#{code.lower()}-{year}-todo)",
            f"- [Election Study Notes and Appendices](#{code.lower()}-{year}-esn)",
            "",
            "---",
            "",
        ]

        # =================================================================
        # Log File Notes
        # =================================================================
        lines.extend([
            f"## Log File Notes: {code}_{year}_M6 {{#{code.lower()}-{year}-log-file-notes}}",
            "",
        ])

        # Deposited Files
        lines.extend(["### Deposited Files", ""])
        df = d.deposited_files

        categories = [
            ("data_files", "Data file(s)"),
            ("questionnaires", "Questionnaire(s)"),
            ("codebooks", "Codebook(s)"),
            ("design_reports", "Design report(s)"),
            ("other", "Other file(s)")
        ]

        for key, label in categories:
            files = df.get(key, [])
            lines.append(f"**{label}:** {len(files)}")
            for f in files:
                lines.append(f"")
                lines.append(f"  - {f}")
            lines.append("")

        # Check for missing required files
        missing = []
        if not df.get("data_files"):
            missing.append("DATA FILE")
        if not df.get("questionnaires"):
            missing.append("QUESTIONNAIRE")
        if not df.get("design_reports"):
            missing.append("DESIGN REPORT")

        if missing:
            lines.append(f"**MISSING:** {', '.join(missing)}")
            lines.append("")

        # Processing Notes
        lines.extend(["### Processing Notes", ""])
        if d.processing_notes:
            for note in d.processing_notes:
                num = note.get("num", 0)
                level = note.get("level", "INFO")
                msg = note.get("message", "")
                prefix = "[!] " if level != "INFO" else ""
                lines.append(f"{num:02d}. {prefix}{msg}")
                lines.append("")
        else:
            lines.append("*No processing notes yet.*")
            lines.append("")

        lines.append("---")
        lines.append("")

        # =================================================================
        # Questions for Collaborator
        # =================================================================
        lines.extend([
            f"## Questions for Collaborator: {code}_{year}_M6 {{#{code.lower()}-{year}-questions}}",
            "",
        ])

        if d.collaborator_questions:
            for q in d.collaborator_questions:
                qid = q.get("id", "?")
                question = q.get("question", "")
                status = q.get("status", "pending").upper()
                step = q.get("step", "?")
                context = q.get("context", "")

                lines.append(f"### {qid} [Step {step}] - {status}")
                lines.append("")
                lines.append(f"{question}")
                if context and context != f"Step {step}":
                    lines.append(f"")
                    lines.append(f"*Context: {context}*")
                lines.append("")
        else:
            lines.append("*No questions for collaborator yet.*")
            lines.append("")

        lines.append("---")
        lines.append("")

        # =================================================================
        # Things To Do Before Releasing
        # =================================================================
        lines.extend([
            f"## Things To Do Before Releasing the Data: {code}_{year}_M6 {{#{code.lower()}-{year}-todo}}",
            "",
        ])

        if d.todo_items:
            for item in d.todo_items:
                lines.append(f"- [ ] {item}")
            lines.append("")
        else:
            lines.append("*No TODO items yet.*")
            lines.append("")

        lines.append("---")
        lines.append("")

        # =================================================================
        # Election Study Notes and Appendices
        # =================================================================
        lines.extend([
            f"## Election Study Notes and Appendices: {code}_{year}_M6 {{#{code.lower()}-{year}-esn}}",
            "",
        ])

        # Election Summary
        lines.extend([f"### Election Summary: {code}_{year}_M6", ""])
        if d.election_summary:
            lines.append(d.election_summary)
            lines.append("")
        else:
            lines.append("*Not yet documented.*")
            lines.append("")

        # Study Design and Weights
        lines.extend([f"### Overview of Study Design and Weights: {code}_{year}_M6", ""])
        sd = d.study_design
        if any(sd.values()):
            field_labels = [
                ("sample_design", "Sample Design"),
                ("sample_size", "Sample Size"),
                ("response_rate", "Response Rate"),
                ("weighting", "Weighting Methodology"),
                ("collection_period", "Data Collection Period"),
                ("mode", "Mode of Interview"),
                ("field_lag", "Field Lag")
            ]
            for key, label in field_labels:
                value = sd.get(key, "") or "TBD"
                lines.append(f"**{label}:** {value}")
                lines.append("")
        else:
            lines.append("*Not yet documented.*")
            lines.append("")

        # Parties and Leaders
        lines.extend([f"### Parties and Leaders: {code}_{year}_M6", ""])
        if d.parties_leaders:
            lines.append(d.parties_leaders)
            lines.append("")
        else:
            lines.append("*Not yet documented.*")
            lines.append("")

        return "\n".join(lines)

    # =========================================================================
    # PUBLIC API - All operations update JSON and re-render Markdown
    # =========================================================================

    def log_message(self, message: str, level: str = "INFO") -> tuple:
        """
        Add a processing note.

        Returns (success, status_message) tuple.
        """
        if not self.log_data:
            if self.state.country and self.state.year:
                self._initialize()
            if not self.log_data:
                return False, "Log not initialized"

        try:
            num = len(self.log_data.processing_notes) + 1
            self.log_data.processing_notes.append({
                "num": num,
                "level": level,
                "message": message,
                "timestamp": datetime.now().isoformat()
            })
            self._save_and_render()
            return True, f"Logged: {message[:50]}..."
        except Exception as e:
            logger.error(f"Failed to log message: {e}")
            return False, f"Error: {e}"

    def update_study_design_section(self, info: dict) -> tuple:
        """
        Update study design fields.

        Args:
            info: Dict with fields to update (sample_design, sample_size, etc.)

        Returns (success, status_message) tuple.
        """
        if not self.log_data:
            return False, "Log not initialized"

        try:
            for key, value in info.items():
                if key in self.log_data.study_design:
                    self.log_data.study_design[key] = value
            self._save_and_render()
            fields = ", ".join(info.keys())
            return True, f"Study design updated: {fields}"
        except Exception as e:
            logger.error(f"Failed to update study design: {e}")
            return False, f"Error: {e}"

    def update_election_summary(self, summary: str) -> tuple:
        """
        Update election summary.

        Returns (success, status_message) tuple.
        """
        if not self.log_data:
            return False, "Log not initialized"

        try:
            self.log_data.election_summary = summary
            self._save_and_render()
            return True, "Election summary updated"
        except Exception as e:
            logger.error(f"Failed to update election summary: {e}")
            return False, f"Error: {e}"

    def update_parties_leaders(self, content: str) -> tuple:
        """
        Update parties and leaders section.

        Returns (success, status_message) tuple.
        """
        if not self.log_data:
            return False, "Log not initialized"

        try:
            self.log_data.parties_leaders = content
            self._save_and_render()
            return True, "Parties and leaders updated"
        except Exception as e:
            logger.error(f"Failed to update parties/leaders: {e}")
            return False, f"Error: {e}"

    def add_collaborator_question(self, question: str, context: str, step_num: int) -> tuple:
        """
        Add a question for the collaborator.

        Returns (success, status_message) tuple.
        """
        if not self.log_data:
            return False, "Log not initialized"

        try:
            # Generate question ID
            num = len(self.log_data.collaborator_questions) + 1
            question_id = f"CQ AA{num}"

            self.log_data.collaborator_questions.append({
                "id": question_id,
                "question": question,
                "context": context,
                "step": step_num,
                "status": "pending",
                "timestamp": datetime.now().isoformat()
            })

            # Also add to state for tracking
            self.state.add_collaborator_question(question, context, step_num)

            self._save_and_render()
            return True, f"Question {question_id} added"
        except Exception as e:
            logger.error(f"Failed to add question: {e}")
            return False, f"Error: {e}"

    def add_todo_item(self, item: str) -> tuple:
        """
        Add a pre-release TODO item.

        Returns (success, status_message) tuple.
        """
        if not self.log_data:
            return False, "Log not initialized"

        try:
            if item not in self.log_data.todo_items:
                self.log_data.todo_items.append(item)
            self._save_and_render()
            return True, f"TODO added: {item[:30]}..."
        except Exception as e:
            logger.error(f"Failed to add TODO: {e}")
            return False, f"Error: {e}"

    def update_deposit_inventory(self, data_files: list, questionnaires: list,
                                  codebooks: list, design_reports: list,
                                  macro_reports: list = None):
        """
        Update the deposited files inventory.
        """
        if not self.log_data:
            return

        try:
            # Convert paths to filenames
            self.log_data.deposited_files = {
                "data_files": [Path(f).name for f in data_files],
                "questionnaires": [Path(f).name for f in questionnaires],
                "codebooks": [Path(f).name for f in codebooks],
                "design_reports": [Path(f).name for f in design_reports],
                "other": [Path(f).name for f in (macro_reports or [])]
            }
            self._save_and_render()
        except Exception as e:
            logger.error(f"Failed to update deposit inventory: {e}")

    # =========================================================================
    # LEGACY API - For compatibility with existing code
    # =========================================================================

    def log_step_start(self, step_num: int, step_name: str):
        """Log the start of a workflow step."""
        self.log_message(f"Step {step_num} started: {step_name}")
        print(f"[Step {step_num}] Starting: {step_name}")

    def log_step_complete(self, step_num: int, message: str, artifacts: list = None):
        """Log the completion of a workflow step."""
        log_msg = f"Step {step_num} completed: {message}"
        if artifacts:
            log_msg += f" (Artifacts: {', '.join(artifacts)})"
        self.log_message(log_msg)
        print(f"[Step {step_num}] Completed: {message}")

    def log_step_issue(self, step_num: int, issue: str, is_question: bool = False):
        """Log an issue encountered during a step."""
        self.log_message(f"Step {step_num} issue: {issue}", level="WARNING")
        print(f"[Step {step_num}] Issue: {issue}")

        if is_question:
            self.add_collaborator_question(issue, f"Step {step_num}", step_num)

    def update_variable_mapping(self, cses_variable: str, source_variable: str, remarks: str = ""):
        """Update the variable tracking sheet (Excel file, not the log)."""
        tracking_file = self.state.variable_tracking_file
        if not tracking_file:
            logger.warning("No variable tracking file set in state")
            return

        tracking_path = Path(tracking_file)
        if not tracking_path.exists():
            logger.warning(f"Variable tracking file not found: {tracking_path}")
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(tracking_path)
            ws = wb['deposited variables']

            found = False
            for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                cell_b = row[1] if len(row) > 1 else None
                if cell_b and cell_b.value and str(cell_b.value).strip() == cses_variable:
                    ws.cell(row=row_idx, column=3, value=source_variable)
                    if remarks:
                        ws.cell(row=row_idx, column=4, value=remarks)
                    found = True
                    break

            if found:
                wb.save(tracking_path)
                print(f"  [Variable] {cses_variable} <- {source_variable}")
            else:
                logger.warning(f"CSES variable {cses_variable} not found in tracking sheet")

        except Exception as e:
            logger.error(f"Failed to update variable tracking sheet: {e}")
